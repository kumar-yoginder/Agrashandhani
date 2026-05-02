import os
import json
import logging
from typing import List, Dict, Any, Tuple
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormatObject
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import Color, Fill
from openpyxl.utils import FORMULAE
from openpyxl.worksheet.table import Table, TableStyleInfo


logger = logging.getLogger(__name__)

MAIN_SHEET_HEADERS = [
    "IOC / Hash / Domain / URL / Apt group / Malware family",
    "IOC Type",
    "VirusTotal (Found/Not Found)",
    "VirusTotal Link",
    "VirusTotal Status",
    "VirusTotal Score",
    "Communicating File MD5 related to the platform",
    "Anyrun (Found/Not Found)",
    "MetaDefender (Found/Not Found)", # Placeholder, not currently in SOURCES
    "Hybrid-Analysis (Found/Not Found)",
    "OPSWAT (Found/Not Found)", # Placeholder, not currently in SOURCES
    "Capesandbox (Found/Not Found)", # Placeholder, not currently in SOURCES
    "Threatfox (Found/Not Found)", # Placeholder, not currently in SOURCES
    "OTX (Found/Not Found)",
    "Timestamp"
]

DETAILED_SHEET_HEADERS = [
    "IOC",
    "IOC Type",
    "Source",
    "Source Query Status",
    "Source Present",
    "Queried At",
    "Raw Data (JSON)",
]

# Define border style for cells
THIN_BORDER = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

def _get_capesandbox_status(vt_result: Dict[str, Any]) -> str:
    """Extracts Capesandbox status from VirusTotal sandbox verdicts.
    
    Returns 'Found' if Capesandbox verdict exists, 'Not Found' otherwise.
    """
    if not vt_result.get("present"):
        return "Not Found"
    
    try:
        vt_api_response = vt_result.get("data", {})
        vt_api_object = vt_api_response.get("data", {})
        vt_api_data = vt_api_object.get("data", {})
        attributes = vt_api_data.get("attributes", {})
        
        # Check for Capesandbox verdict in sandbox_verdicts
        sandbox_verdicts = attributes.get("sandbox_verdicts", {})
        
        # Look for Capesandbox in the sandbox verdicts
        if "Capesandbox" in sandbox_verdicts:
            capesandbox_verdict = sandbox_verdicts.get("Capesandbox", {})
            if isinstance(capesandbox_verdict, dict) and capesandbox_verdict.get("category"):
                return "Found"
    except Exception as e:
        logger.debug(f"Error extracting Capesandbox status: {e}")
    
    return "Not Found"

def _get_virustotal_link_and_status(vt_result: Dict[str, Any], ioc_value: str, ioc_type_classified: str) -> Tuple[str, str, str]:
    """Extracts VirusTotal link, status, and score from VT result.
    
    Status values:
    - 'malicious': Multiple AV engines detect as malicious
    - 'malware': Sandbox verdicts or threat classification indicate malware
    - 'clean': No detections or only harmless detections
    - 'unknown': Unclear verdict (e.g., only timeout or undetected)
    """
    link = ""
    status = "not found"
    score = ""

    if vt_result.get("present") and "data" in vt_result.get("data", {}):
        # Navigate through nested data structure:
        # vt_result["data"] = {query_status, source, data}
        # vt_result["data"]["data"] = {data: {actual_vt_response}}
        # vt_result["data"]["data"]["data"] = {id, type, attributes, ...}
        vt_api_response = vt_result.get("data", {})
        vt_api_object = vt_api_response.get("data", {})
        vt_api_data = vt_api_object.get("data", {})
        attributes = vt_api_data.get("attributes", {})
        
        # Link construction based on the classified IOC type
        if "ip" in ioc_type_classified.lower():
            link = f"https://www.virustotal.com/gui/ip-address/{ioc_value}"
        elif "domain" in ioc_type_classified.lower():
            link = f"https://www.virustotal.com/gui/domain/{ioc_value}"
        elif "hash" in ioc_type_classified.lower():
            link = f"https://www.virustotal.com/gui/file/{ioc_value}"
        elif "url" in ioc_type_classified.lower():
            try:
                import base64
                url_id = base64.urlsafe_b64encode(ioc_value.encode()).decode().rstrip("=")
                link = f"https://www.virustotal.com/gui/url/{url_id}"
            except Exception:
                link = f"https://www.virustotal.com/gui/search/{ioc_value}"

        # Extract analysis statistics
        last_analysis_stats = attributes.get("last_analysis_stats", {})
        harmless = last_analysis_stats.get("harmless", 0)
        malicious = last_analysis_stats.get("malicious", 0)
        suspicious = last_analysis_stats.get("suspicious", 0)
        undetected = last_analysis_stats.get("undetected", 0)
        
        total = harmless + malicious + suspicious + undetected
        
        # Calculate score
        if total > 0:
            score = f"{malicious}/{total}"
        else:
            score = "0/0"
            status = "unknown"
            return link, status, score
        
        # Determine status - priority-based logic
        # 1. Check sandbox verdicts for malware confirmation
        sandbox_verdicts = attributes.get("sandbox_verdicts", {})
        malware_detected_in_sandbox = False
        for sandbox_name, verdict_data in sandbox_verdicts.items():
            if isinstance(verdict_data, dict):
                category = verdict_data.get("category", "").lower()
                malware_class = verdict_data.get("malware_classification", [])
                if "malicious" in category or "MALWARE" in malware_class:
                    malware_detected_in_sandbox = True
                    break
        
        # 2. Check popular threat classification
        threat_classification = attributes.get("popular_threat_classification", {})
        suggested_label = threat_classification.get("suggested_threat_label", "").lower()
        threat_categories = threat_classification.get("popular_threat_category", [])
        has_malware_classification = any(
            cat.get("value", "").lower() in ["trojan", "malware", "backdoor", "worm", "ransomware", "spyware"]
            for cat in threat_categories if isinstance(cat, dict)
        )
        
        # 3. Determine final status
        if malicious > 0:
            status = "malicious"
        elif malware_detected_in_sandbox or has_malware_classification or "malware" in suggested_label:
            status = "malware"
        elif harmless > 0 and malicious == 0 and suspicious == 0:
            status = "clean"
        elif suspicious > 0:
            status = "malware"  # Treat suspicious as potential malware
        else:
            status = "unknown"  # Default for undetected or timeouts only
        
    return link, status, score

def _get_source_status_string(source_result: Dict[str, Any]) -> str:
    """Returns 'Found', 'Not Found', or 'Error' based on source result."""
    if source_result.get("present"):
        return "Found"
    elif source_result.get("data", {}).get("query_status") == "not_found":
        return "Not Found"
    elif source_result.get("data", {}).get("query_status") == "error":
        return "Error"
    return "Not Found" # Default for absence or unclear status


def _extract_main_sheet_data(result: Dict[str, Any]) -> Dict[str, Any]:
    """Extracts and formats data for the main sheet from a single IOC result."""
    main_data = {header: "" for header in MAIN_SHEET_HEADERS}
    ioc_value = result.get("query", "N/A")

    main_data["IOC / Hash / Domain / URL / Apt group / Malware family"] = ioc_value
    main_data["IOC Type"] = result.get("ioc_type", "unknown").replace("_", " ").title()
    main_data["Timestamp"] = result.get("timestamp", "")

    sources_results = result.get("sources", {})
    ioc_type_classified = result.get("ioc_type", "unknown")

    # VirusTotal
    vt_result = sources_results.get("virustotal", {})
    if vt_result:
        main_data["VirusTotal (Found/Not Found)"] = _get_source_status_string(vt_result)
        vt_link, vt_status, vt_score = _get_virustotal_link_and_status(vt_result, ioc_value, ioc_type_classified)
        main_data["VirusTotal Link"] = vt_link
        main_data["VirusTotal Status"] = vt_status
        main_data["VirusTotal Score"] = vt_score

    # Communicating File MD5 (from OTX/VT relations)
    otx_result = sources_results.get("otx", {})
    communicating_md5s = []
    if otx_result.get("present"):
        otx_api_response = otx_result.get("data", {})  # {query_status, source, data}
        otx_api_data = otx_api_response.get("data", {})  # The actual OTX response
        correlated_files = otx_api_data.get("correlated_hashes", {}).get("related_files", [])
        # OTX's related_files are already hashes (MD5, SHA1, SHA256)
        if correlated_files:
            # Limit to a reasonable number to avoid overflowing cells
            md5s_to_display = [h for h in correlated_files if len(h) == 32] # Filter for MD5s if possible
            if md5s_to_display:
                communicating_md5s.extend(md5s_to_display)
            else:
                # If no MD5s specifically, take first few available hashes
                communicating_md5s.extend(correlated_files)
            
            if len(communicating_md5s) > 3:
                main_data["Communicating File MD5 related to the platform"] = "; ".join(communicating_md5s[:3]) + f" (and {len(communicating_md5s) - 3} more)"
            elif communicating_md5s:
                main_data["Communicating File MD5 related to the platform"] = "; ".join(communicating_md5s)
    
    # If still empty, check VirusTotal for communicating files (more complex, often requires parsing specific VT reports)
    # For simplicity, if OTX found anything, we prioritize that. Otherwise, mark as not found.
    if not communicating_md5s:
        main_data["Communicating File MD5 related to the platform"] = "Not Found"


    # Capesandbox (from VirusTotal sandbox verdicts)
    main_data["Capesandbox (Found/Not Found)"] = _get_capesandbox_status(vt_result)

    # Other sources
    source_mappings = {
        "anyrun": "Anyrun (Found/Not Found)",
        "hybrid_analysis": "Hybrid-Analysis (Found/Not Found)",
        "otx": "OTX (Found/Not Found)",
        # Placeholders for sources not yet integrated or providing specific status
        # "metadefender": "MetaDefender (Found/Not Found)",
        # "opswat": "OPSWAT (Found/Not Found)",
        # "threatfox": "Threatfox (Found/Not Found)",
    }
    for source_name, header in source_mappings.items():
        if source_name in sources_results:
            main_data[header] = _get_source_status_string(sources_results[source_name])
        else:
            main_data[header] = "N/A (Source not queried or available)"


    return main_data

def _extract_detailed_sheet_data(result: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Extracts and formats data for the detailed sheet from a single IOC result."""
    detailed_rows = []
    ioc_value = result.get("query", "N/A")
    ioc_type = result.get("ioc_type", "unknown")
    queried_at_global = result.get("timestamp", "")

    for source_name, source_result in result.get("sources", {}).items():
        raw_data = source_result.get("data", {})
        detailed_rows.append({
            "IOC": ioc_value,
            "IOC Type": ioc_type,
            "Source": source_name,
            "Source Query Status": raw_data.get("query_status", "N/A"),
            "Source Present": source_result.get("present", False),
            "Queried At": source_result.get("queried_at", queried_at_global),
            "Raw Data (JSON)": json.dumps(raw_data, indent=2)
        })
    return detailed_rows

def _apply_styles_and_width(ws, headers):
    """Applies basic styling and sets column widths."""
    # Header styling
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid") # Blue
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # Set column widths based on header length or content heuristic
    for col_idx, header in enumerate(headers, 1):
        # A simple heuristic for column width
        if "JSON" in header:
            ws.column_dimensions[get_column_letter(col_idx)].width = 80
        elif "Link" in header:
             ws.column_dimensions[get_column_letter(col_idx)].width = 50
        else:
            ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header), 15) + 2 # Min width 15, plus padding

    # Apply data styles (e.g., borders)
    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).border = THIN_BORDER

    # Conditional Formatting for "Main" sheet status
    if ws.title == "Main":
        # Green for Found/Clean
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Light Green
        for col_idx, header in enumerate(headers, 1):
            if "Found" in header or "Status" in header:
                # Apply to range of data cells (start from row 2)
                ws.conditional_formatting.add(
                    f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{ws.max_row}",
                    CellIsRule(operator='equal', formula=['"Found"'], fill=green_fill)
                )
                ws.conditional_formatting.add(
                    f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{ws.max_row}",
                    CellIsRule(operator='equal', formula=['"clean"'], fill=green_fill)
                )

        # Red for Malicious/Malware/Error
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Light Red
        for col_idx, header in enumerate(headers, 1):
            if "Found" in header or "Status" in header:
                ws.conditional_formatting.add(
                    f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{ws.max_row}",
                    CellIsRule(operator='equal', formula=['"malicious"'], fill=red_fill)
                )
                ws.conditional_formatting.add(
                    f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{ws.max_row}",
                    CellIsRule(operator='equal', formula=['"malware"'], fill=red_fill)
                )
                ws.conditional_formatting.add(
                    f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{ws.max_row}",
                    CellIsRule(operator='equal', formula=['"Error"'], fill=red_fill)
                )
        
        # Yellow for Unknown/Not Found
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Light Yellow
        for col_idx, header in enumerate(headers, 1):
            if "Found" in header or "Status" in header:
                ws.conditional_formatting.add(
                    f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{ws.max_row}",
                    CellIsRule(operator='equal', formula=['"Not Found"'], fill=yellow_fill)
                )
                ws.conditional_formatting.add(
                    f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{ws.max_row}",
                    CellIsRule(operator='equal', formula=['"unknown"'], fill=yellow_fill)
                )


def export_to_excel(results_list: List[Dict[str, Any]], output_filepath: str, update_existing: bool = False):
    """
    Exports OSINT search results to an Excel file with 'Main' and 'Detailed' sheets.
    Can create a new file or update an existing one.
    """
    if update_existing and os.path.exists(output_filepath):
        try:
            workbook = load_workbook(output_filepath)
            logger.info(f"Loaded existing Excel workbook: {output_filepath}")
        except Exception as e:
            logger.error(f"Error loading existing Excel file {output_filepath}, creating new one. Error: {e}")
            workbook = Workbook()
            
    else:
        workbook = Workbook()
        
    # Get or create Main sheet
    if "Main" in workbook.sheetnames:
        main_ws = workbook["Main"]
        # Clear existing data if not updating or if first row is not headers
        if not update_existing or main_ws.cell(row=1, column=1).value != MAIN_SHEET_HEADERS[0]:
            workbook.remove(main_ws)
            main_ws = workbook.create_sheet("Main", 0)
            main_ws.append(MAIN_SHEET_HEADERS) # Add headers only if new sheet
    else:
        main_ws = workbook.create_sheet("Main", 0)
        main_ws.append(MAIN_SHEET_HEADERS) # Add headers
    
    # Get or create Detailed sheet
    if "Detailed" in workbook.sheetnames:
        detailed_ws = workbook["Detailed"]
        if not update_existing or detailed_ws.cell(row=1, column=1).value != DETAILED_SHEET_HEADERS[0]:
            workbook.remove(detailed_ws)
            detailed_ws = workbook.create_sheet("Detailed", 1)
            detailed_ws.append(DETAILED_SHEET_HEADERS) # Add headers only if new sheet
    else:
        detailed_ws = workbook.create_sheet("Detailed", 1)
        detailed_ws.append(DETAILED_SHEET_HEADERS) # Add headers

    # Remove default "Sheet" if it's empty
    # Remove default "Sheet" if it's empty and only contains a single default cell
    if "Sheet" in workbook.sheetnames:
        default_sheet = workbook["Sheet"]
        # Check if the default sheet is effectively empty (only initial cell or completely empty)
        if default_sheet.max_row == 1 and default_sheet.max_column == 1 and default_sheet.cell(1, 1).value is None:
            workbook.remove(default_sheet)

    existing_iocs_main = set()
    existing_iocs_detailed = set()

    if update_existing:
        # Read existing IOCs from Main sheet
        for row_idx in range(2, main_ws.max_row + 1): # Skip header row
            ioc_value = main_ws.cell(row=row_idx, column=1).value
            if ioc_value:
                existing_iocs_main.add(ioc_value)
        
        # Read existing IOCs/Source combinations from Detailed sheet
        for row_idx in range(2, detailed_ws.max_row + 1): # Skip header row
            ioc_value = detailed_ws.cell(row=row_idx, column=1).value
            source_name = detailed_ws.cell(row=row_idx, column=3).value
            if ioc_value and source_name:
                existing_iocs_detailed.add(f"{ioc_value}-{source_name}")

    main_rows_to_write = []
    detailed_rows_to_write = []

    for result in results_list:
        ioc_value = result.get("query", "N/A")

        # Process for Main sheet
        if ioc_value not in existing_iocs_main or not update_existing:
            main_rows_to_write.append(_extract_main_sheet_data(result))
        elif update_existing:
            # If updating, find the row and update it
            for row_idx in range(2, main_ws.max_row + 1):
                if main_ws.cell(row=row_idx, column=1).value == ioc_value:
                    updated_data = _extract_main_sheet_data(result)
                    for col_idx, header in enumerate(MAIN_SHEET_HEADERS, 1):
                        main_ws.cell(row=row_idx, column=col_idx, value=updated_data.get(header, ""))
                    logger.debug(f"Updated IOC '{ioc_value}' in Main sheet.")
                    break
        
        # Process for Detailed sheet
        for detailed_row_data in _extract_detailed_sheet_data(result):
            source_name = detailed_row_data.get("Source", "N/A")
            composite_key = f"{ioc_value}-{source_name}"
            if composite_key not in existing_iocs_detailed or not update_existing:
                detailed_rows_to_write.append(detailed_row_data)
            elif update_existing:
                # Find the row and update it
                for row_idx in range(2, detailed_ws.max_row + 1):
                    existing_ioc = detailed_ws.cell(row=row_idx, column=1).value
                    existing_source = detailed_ws.cell(row=row_idx, column=3).value
                    if existing_ioc == ioc_value and existing_source == source_name:
                        for col_idx, header in enumerate(DETAILED_SHEET_HEADERS, 1):
                            detailed_ws.cell(row=row_idx, column=col_idx, value=detailed_row_data.get(header, ""))
                        logger.debug(f"Updated IOC '{ioc_value}' for source '{source_name}' in Detailed sheet.")
                        break

    # Append new rows
    for row_data in main_rows_to_write:
        main_ws.append([row_data.get(header, "") for header in MAIN_SHEET_HEADERS])
    
    for row_data in detailed_rows_to_write:
        detailed_ws.append([row_data.get(header, "") for header in DETAILED_SHEET_HEADERS])

    # Apply styles and set column widths
    _apply_styles_and_width(main_ws, MAIN_SHEET_HEADERS)
    _apply_styles_and_width(detailed_ws, DETAILED_SHEET_HEADERS)
    
    try:
        workbook.save(output_filepath)
        logger.info(f"Results successfully exported to Excel: {output_filepath}")
        return output_filepath
    except Exception as e:
        logger.error(f"Error saving Excel file {output_filepath}: {e}")
        return None
